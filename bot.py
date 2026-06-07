"""
================================================================
  BOT TELEGRAM — RAPPORT JOURNALIER IDCAM
  Technicien : ABDOUL-AZIZ MOHAMADOU LAMINOU
================================================================
"""

import os
import datetime
import random
import warnings
import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, ContextTypes
)
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore")
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# ─────────────────────────────────────────────
#  ⚙️  CONFIGURATION — MODIFIEZ ICI
# ─────────────────────────────────────────────

TOKEN_BOT      = "8871644850:AAGqLgYPVo8ABS0xFtAhJJDoaa3_XdiN7Sk"   # ← Token donné par @BotFather
TEMPLATE_PATH  = os.path.join(os.path.dirname(__file__), "template_rapport.xlsx")
DOSSIER_SORTIE = os.path.dirname(__file__)

# ─────────────────────────────────────────────
#  DONNÉES : INTERVENTIONS
# ─────────────────────────────────────────────

POSTES_LOCAL     = [f"poste {i:02d}" for i in range(1, 21)]
CENTRES_DISTANCE = ["NO27 DJAMBOUTOU", "NO29 YELWA", "NO06 LAGO", "NO03 FIGUIL", "NO08 GUIDER"]

POOL = [
    ("Logiciel","Enrôlement","local",
     "camera scan doc - ecran noir",
     lambda p,c: f"Le {p} (centre {c}): camera scan doc affiche ecran noir. Reinitialisation camera et redemarrage machine. Resolu."),
    ("Logiciel","Enrôlement","local",
     "camera figee sur ecran blanc",
     lambda p,c: f"Camera du {p} figee sur ecran blanc. Redemarrage complet de l'ordinateur. Resolu."),
    ("Logiciel","Enrôlement","local",
     "probleme reseau - camera et empreinte HS",
     lambda p,c: f"Le {p}: panne reseau rendant camera et empreinte inactives. Verification cable, test ping, redemarrage. Resolu."),
    ("Logiciel","Enrôlement","local",
     "empreinte digitale non fonctionnelle",
     lambda p,c: f"Le {p}: empreinte digitale dysfonctionnelle. Debranchement/rebranchement lecteur biometrique et redemarrage. Resolu."),
    ("Logiciel","Enrôlement","local",
     "QR code non detecte par camera doc",
     lambda p,c: f"Le {p}: camera scan doc ne detecte pas le QR code. Retelechargement document et nouvelle tentative. Resolu."),
    ("Logiciel","Enrôlement","local",
     "ajustement photo - personne agee ou handicapee",
     lambda p,c: f"Le {p}: difficulte prise de photo usager age. Assistance enrolleur pour reglage luminosite et cadrage. Photo validee."),
    ("Logiciel","Enrôlement","local",
     "logiciel QOMITTO ne s'ouvre pas",
     lambda p,c: f"Le {p}: QOMITTO ne s'ouvre pas. Redemarrage du service et relance application. Resolu."),
    ("Logiciel","Enrôlement","local",
     "redemarrage correctif - poste lent",
     lambda p,c: f"Le {p} presentait des lenteurs. Redemarrage complet effectue. Performances retablies."),
    ("Logiciel","Enrôlement","local",
     "empreinte et camera inactives apres coupure reseau",
     lambda p,c: f"Suite a coupure reseau sur {p}: empreinte et camera inactives. Redemarrage apres retablissement reseau. Resolu."),
    ("Logiciel","Enrôlement","distance",
     "assistance distance - probleme reseau",
     lambda p,c: f"Assistance telephonique centre {c}: probleme reseau. Guidage test ping et redemarrage. Resolu a distance."),
    ("Logiciel","Enrôlement","distance",
     "assistance distance - empreinte non reconnue",
     lambda p,c: f"Assistance a distance centre {c}: empreinte non reconnue. Instructions debranchement/rebranchement. Resolu."),
    ("Logiciel","Enrôlement","distance",
     "assistance distance - camera bloquee",
     lambda p,c: f"Assistance telephonique centre {c}: camera bloquee. Guidage reinitialisation camera. Resolu a distance."),
    ("Logiciel","Enrôlement","distance",
     "assistance distance - QOMITTO ne fonctionne pas",
     lambda p,c: f"Assistance a distance centre {c}: QOMITTO ne fonctionne pas. Guidage redemarrage service. Resolu."),
]

# ─────────────────────────────────────────────
#  GÉNÉRATION DU RAPPORT
# ─────────────────────────────────────────────

def pick_interventions(n):
    ivs, seen = [], set()
    for _ in range(n * 10):
        if len(ivs) >= n: break
        typ, ss, loc, titre, desc_fn = random.choice(POOL)
        if titre in seen: continue
        seen.add(titre)
        if loc == "distance":
            centre = random.choice(CENTRES_DISTANCE)
            code, desc = centre, desc_fn("", centre)
        else:
            poste = random.choice(POSTES_LOCAL)
            code, desc = "NO01BIS", desc_fn(poste, "NO01BIS")
        ivs.append({"type": typ, "ss": ss, "code": code, "titre": titre, "desc": desc})
    return ivs

def rand_times(n):
    slots, used = [], set()
    for _ in range(n * 20):
        if len(slots) >= n: break
        sm = random.randint(8*60+30, 15*60)
        k = sm // 15
        if k in used: continue
        used.add(k)
        em = min(sm + random.randint(10, 45), 16*60)
        slots.append((sm/1440, em/1440))
    return sorted(slots)

def frac_to_time(frac):
    s = int(round(frac * 86400))
    h, r = divmod(s, 3600); m, s = divmod(r, 60)
    return datetime.time(h, m, s)

def generer_rapport(target_date):
    random.seed(int(target_date.strftime("%Y%m%d")))
    n    = random.randint(3, 6)
    ivs  = pick_interventions(n)
    tms  = rand_times(len(ivs))
    tot  = sum(e - s for s, e in tms)
    n_log = sum(1 for iv in ivs if iv["type"] == "Logiciel")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Rapport Journalier"]

    ws["B2"] = target_date
    ws["B2"].number_format = "dd/mm/yyyy"

    for r in range(6, 9):
        for col in "CDEFG":
            ws[f"{col}{r}"] = None
    ws["D6"] = n_log

    ws["B13"] = "ABDOUL-AZIZ M.L"
    dv = DataValidation(type="list", formula1="Liste!$B$2:$B$34",
                        allow_blank=False, showDropDown=False)
    dv.sqref = "B13"
    ws.add_data_validation(dv)

    ws["C13"] = len(ivs)
    ws["D13"] = frac_to_time(tot)
    ws["D13"].number_format = "h:mm:ss"

    for r in range(18, 24):
        for col in "ABCDEFGHIJK":
            ws[f"{col}{r}"].value = None

    tpl_rows = 4
    if len(ivs) > tpl_rows:
        ws.insert_rows(22, len(ivs) - tpl_rows)
    elif len(ivs) < tpl_rows:
        ws.delete_rows(18 + len(ivs), tpl_rows - len(ivs))

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    dv2 = DataValidation(type="list", formula1="Liste!$B$2:$B$34",
                         allow_blank=False, showDropDown=False)
    dv2.sqref = f"B18:B{17+len(ivs)}"
    ws.add_data_validation(dv2)

    for i, (iv, (ts, te)) in enumerate(zip(ivs, tms)):
        r = 18 + i
        def sc(col, val, fmt=None, center=False):
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.font = Font(name="Calibri", size=11)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if center else "left",
                vertical="center" if center else "top",
                wrap_text=True
            )
            if fmt: cell.number_format = fmt

        sc("A", i+1, center=True)
        sc("B", "ABDOUL-AZIZ M.L")
        sc("C", iv["type"])
        sc("D", iv["ss"])
        sc("E", iv["code"], center=True)
        sc("F", iv["titre"])
        sc("G", iv["desc"])
        sc("H", "", center=True)
        sc("I", frac_to_time(ts), fmt="h:mm", center=True)
        sc("J", "Terminé", center=True)
        sc("K", frac_to_time(te), fmt="h:mm", center=True)
        ws.row_dimensions[r].height = 57.0

    nom = f"Rapport_Journalier_ABDOUL-AZIZ_{target_date.strftime('%Y-%m-%d')}.xlsx"
    chemin = os.path.join(DOSSIER_SORTIE, nom)
    wb.save(chemin)
    return chemin, ivs, tms, frac_to_time(tot)

# ─────────────────────────────────────────────
#  COMMANDES DU BOT
# ─────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Commande /start — message de bienvenue."""
    texte = (
        "👋 *Bonjour ABDOUL-AZIZ !*\n\n"
        "Je suis votre assistant IDCAM. Je génère automatiquement "
        "votre rapport journalier Excel.\n\n"
        "📋 *Commandes disponibles :*\n"
        "• /rapport — Générer le rapport d'aujourd'hui\n"
        "• /rapport\\_hier — Générer le rapport d'hier\n"
        "• /aide — Afficher l'aide\n\n"
        "Ou utilisez les boutons ci-dessous 👇"
    )
    clavier = InlineKeyboardMarkup([
        [InlineKeyboardButton("📄 Rapport aujourd'hui", callback_data="rapport_today")],
        [InlineKeyboardButton("📅 Rapport hier",        callback_data="rapport_hier")],
        [InlineKeyboardButton("❓ Aide",                callback_data="aide")],
    ])
    await update.message.reply_text(texte, parse_mode="Markdown", reply_markup=clavier)


async def cmd_rapport(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Commande /rapport — génère le rapport du jour."""
    await _envoyer_rapport(update, datetime.date.today())


async def cmd_rapport_hier(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Commande /rapport_hier — génère le rapport d'hier."""
    hier = datetime.date.today() - datetime.timedelta(days=1)
    await _envoyer_rapport(update, hier)


async def cmd_aide(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📖 *AIDE — Bot Rapport IDCAM*\n\n"
        "• `/rapport` → Rapport d'aujourd'hui\n"
        "• `/rapport_hier` → Rapport d'hier\n"
        "• `/start` → Menu principal\n\n"
        "Le fichier Excel généré contient :\n"
        "✅ La date automatique\n"
        "✅ Vos interventions du jour\n"
        "✅ Les formules et couleurs d'origine\n"
        "✅ La liste déroulante des techniciens\n"
        "✅ La durée totale calculée\n\n"
        "📞 Centre : NO01BIS + assistance à distance",
        parse_mode="Markdown"
    )


async def bouton(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Gestion des boutons inline."""
    query = update.callback_query
    await query.answer()
    if query.data == "rapport_today":
        await _envoyer_rapport(query, datetime.date.today())
    elif query.data == "rapport_hier":
        hier = datetime.date.today() - datetime.timedelta(days=1)
        await _envoyer_rapport(query, hier)
    elif query.data == "aide":
        await query.message.reply_text(
            "📖 Commandes : /rapport | /rapport\\_hier | /start",
            parse_mode="Markdown"
        )


async def _envoyer_rapport(cible, date):
    """Génère et envoie le fichier Excel."""
    # Message d'attente
    if hasattr(cible, "message"):
        msg = await cible.message.reply_text("⏳ Génération du rapport en cours...")
    else:
        msg = await cible.reply_text("⏳ Génération du rapport en cours...")

    try:
        chemin, ivs, tms, duree = generer_rapport(date)

        # Résumé texte
        resume = (
            f"✅ *Rapport du {date.strftime('%d/%m/%Y')} généré !*\n\n"
            f"👤 Technicien : ABDOUL-AZIZ M.L\n"
            f"🏢 Centre : NO01BIS\n"
            f"📊 Interventions : *{len(ivs)}*\n"
            f"⏱ Durée totale : *{duree}*\n\n"
            f"*Détail :*\n"
        )
        for i, iv in enumerate(ivs):
            resume += f"  {i+1}. [{iv['code']}] {iv['titre']}\n"

        await msg.delete()

        # Envoyer fichier Excel
        if hasattr(cible, "message"):
            await cible.message.reply_document(
                document=open(chemin, "rb"),
                filename=os.path.basename(chemin),
                caption=resume,
                parse_mode="Markdown"
            )
        else:
            await cible.reply_document(
                document=open(chemin, "rb"),
                filename=os.path.basename(chemin),
                caption=resume,
                parse_mode="Markdown"
            )

        # Nettoyer le fichier local
        os.remove(chemin)

    except FileNotFoundError:
        await msg.edit_text(
            "❌ Fichier template introuvable.\n"
            "Vérifiez que `template_rapport.xlsx` est dans le même dossier que `bot.py`."
        )
    except Exception as e:
        await msg.edit_text(f"❌ Erreur : {str(e)}")


# ─────────────────────────────────────────────
#  ENVOI AUTOMATIQUE QUOTIDIEN À 15H
# ─────────────────────────────────────────────

async def rapport_auto(context: ContextTypes.DEFAULT_TYPE):
    """Tâche planifiée : envoie le rapport automatiquement à 15h."""
    chat_id = context.job.chat_id
    date = datetime.date.today()
    try:
        chemin, ivs, tms, duree = generer_rapport(date)
        texte = (
            f"🕒 *Rapport automatique 15h00*\n"
            f"📅 {date.strftime('%d/%m/%Y')}\n"
            f"👤 ABDOUL-AZIZ M.L | Centre NO01BIS\n"
            f"📊 {len(ivs)} interventions | ⏱ {duree}"
        )
        await context.bot.send_document(
            chat_id=chat_id,
            document=open(chemin, "rb"),
            filename=os.path.basename(chemin),
            caption=texte,
            parse_mode="Markdown"
        )
        os.remove(chemin)
    except Exception as e:
        await context.bot.send_message(chat_id=chat_id, text=f"❌ Erreur rapport auto : {e}")


async def activer_auto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Commande /auto — active l'envoi automatique à 15h00."""
    chat_id = update.effective_chat.id
    # Supprimer ancienne tâche si existe
    current_jobs = context.job_queue.get_jobs_by_name("rapport_auto")
    for job in current_jobs:
        job.schedule_removal()

    # Planifier à 15h00 chaque jour (heure du serveur = UTC, ajuster si besoin)
    context.job_queue.run_daily(
        rapport_auto,
        time=datetime.time(hour=14, minute=0, second=0),  # 14h UTC = 15h Cameroun (UTC+1)
        chat_id=chat_id,
        name="rapport_auto"
    )
    await update.message.reply_text(
        "✅ *Envoi automatique activé !*\n"
        "Je vous enverrai votre rapport chaque jour à *15h00* automatiquement.\n\n"
        "Pour désactiver : /stopauto",
        parse_mode="Markdown"
    )


async def stopper_auto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Commande /stopauto — désactive l'envoi automatique."""
    jobs = context.job_queue.get_jobs_by_name("rapport_auto")
    for job in jobs:
        job.schedule_removal()
    await update.message.reply_text("🔕 Envoi automatique désactivé.")


# ─────────────────────────────────────────────
#  LANCEMENT DU BOT
# ─────────────────────────────────────────────

def main():
    print("🤖 Bot IDCAM démarré...")
    app = Application.builder().token(TOKEN_BOT).build()

    app.add_handler(CommandHandler("start",         start))
    app.add_handler(CommandHandler("rapport",       cmd_rapport))
    app.add_handler(CommandHandler("rapport_hier",  cmd_rapport_hier))
    app.add_handler(CommandHandler("aide",          cmd_aide))
    app.add_handler(CommandHandler("auto",          activer_auto))
    app.add_handler(CommandHandler("stopauto",      stopper_auto))
    app.add_handler(CallbackQueryHandler(bouton))

    print("✅ Bot en écoute. Appuyez sur Ctrl+C pour arrêter.")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
